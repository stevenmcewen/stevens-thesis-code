import openpyxl
import os


class Excel:

    def __init__(self, folder_path, output_filename):
        self.folder_path = folder_path
        self.output_filename = output_filename
        self.files = os.listdir(self.folder_path)
        self.workbook = openpyxl.Workbook()
        self.workbook.remove_sheet(self.workbook.active)

    def combine_sheets(self):
        print("Combining sheets from Excel files...")
        desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')
        file_path = os.path.join(desktop_path, self.output_filename)
        if os.path.exists(file_path):
            # File already exists, delete it
            pass
        else:
            for file in self.files:
                if file.endswith('.xlsx'):
                    file_name, file_ext = os.path.splitext(file)
                    wb = openpyxl.load_workbook(os.path.join(self.folder_path, file))
                    for sheet in wb:
                        new_sheet = self.workbook.create_sheet(title=file_name)
                        new_sheet.insert_rows(0)
                        for row in sheet.rows:
                            for cell in row:
                                new_sheet[cell.coordinate].value = cell.value
            self.workbook.save(file_path)


