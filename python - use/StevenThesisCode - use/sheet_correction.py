import openpyxl
import math

conversion_dict = {
    '1EP': "1EP", '2EP': "1.5EP", '3EP': "2EP", '4EP': "3EP", '5EP': "4EP", '6EP': "4.5EP", '7EP': "5.5EP",
    '8EP': "6EP",
    '9EP': "7EP", '10EP': "7.5EP", '11EP': "8.5EP", '12EP': "9EP", '13EP': "10EP", '14EP': "10.5EP", '15EP': "11.5EP",
    '16EP': "12EP", '17EP': "13EP", '18EP': "14EP", '19EP': "14.5EP", '20EP': "15EP", '21EP': "16EP", '22EP': "17EP",
    '23EP': "17.5EP", '24EP': "18EP", '25EP': "19EP", '26EP': "19.5EP", '27EP': "20.5EP", '28EP': "21EP",
    '29EP': "22EP",
    '30EP': "22.5EP", '31EP': "23EP", '32EP': "24EP", '33EP': "24.5EP", '34EP': "25EP", '35EP': "26EP",
    '36EP': "26.5EP",
    '37EP': "27.5EP", '38EP': "28EP", '39EP': "28.5EP", '40EP': "29.5EP", '41EP': "30EP", '42EP': "30.5EP",
    '43EP': "31EP", '44EP': "32EP", '45EP': "32.5EP", '46EP': "33EP", '47EP': "34EP", '48EP': "34.5EP", '49EP': "35EP",
    '50EP': "36EP", '51EP': "36.5EP", '52EP': "37EP", '53EP': "37.5EP", '54EP': "38EP", '55EP': "38.5EP",
    '56EP': "39EP",
    '57EP': "40EP", '58EP': "40.5EP", '59EP': "41EP", '60EP': "41.5EP", '61EP': "42EP", '62EP': "42.5EP",
    '63EP': "43EP",
    '64EP': "43.5EP", '65EP': "44EP", '66EP': "44EP", '67EP': "44.5EP", '68EP': "45EP", '69EP': "45.5EP",
    '70EP': "46EP",
    '71EP': "46EP", '72EP': "46.5EP", '73EP': "46.5EP", '74EP': "47EP", '75EP': "47.5EP", '1EE': "1EE", '2EE': "2EE",
    '3EE': "3EE", '4EE': "4EE", '5EE': "6EE", '6EE': "7EE", '7EE': "8EE", '8EE': "9EE", '9EE': "10EE", '10EE': "11EE",
    '11EE': "13EE", '12EE': "14EE", '13EE': "15EE", '14EE': "16EE", '15EE': "17EE", '16EE': "18EE", '17EE': "19EE",
    '18EE': "20EE", '19EE': "22EE", '20EE': "23EE", '21EE': "24EE", '22EE': "25EE", '23EE': "26EE", '24EE': "27EE",
    '25EE': "28EE",
    '26EE': "29EE", '27EE': "30EE", '28EE': "32EE", '29EE': "33EE", '30EE': "34EE", '31EE': "35EE", '32EE': "36EE",
    '33EE': "37EE",
    '34EE': "38EE", '35EE': "39EE", '36EE': "41EE", '37EE': "42EE", '38EE': "43EE", '39EE': "44EE", '40EE': "45EE",
    '41EE': "46EE", '42EE': "47EE",
    '43EE': "48EE", '44EE': "50EE", '45EE': "51EE", '46EE': "52EE", '47EE': "53EE", '48EE': "54EE", '49EE': "55EE",
    '50EE': "56EE",
    '51EE': "58EE", '52EE': "59EE", '53EE': "60EE", '54EE': "61EE", '55EE': "62EE", '56EE': "63EE", '57EE': "64EE",
    '58EE': "65EE",
    '1WP': "1WP", '2WP': "1.5WP", '3WP': "2WP", '4WP': "3WP", '5WP': "4WP", '6WP': "4.5WP", '7WP': "5.5WP",
    '8WP': "6WP", '9WP': "7WP", '10WP': "7.5WP", '11WP': "8.5WP", '12WP': "9WP", '13WP': "10WP", '14WP': "10.5WP",
    '15WP': "11.5WP", '16WP': "12WP", '17WP': "13WP", '18WP': "14WP", '19WP': "14.5WP", '20WP': "15WP", '21WP': "16WP",
    '22WP': "17WP", '23WP': "17.5WP", '24WP': "18WP", '25WP': "19WP", '26WP': "19.5WP", '27WP': "20.5WP",
    '28WP': "21WP",
    '29WP': "22WP", '30WP': "22.5WP", '31WP': "23WP", '32WP': "24WP", '33WP': "24.5WP", '34WP': "25WP", '35WP': "26WP",
    '36WP': "26.5WP", '37WP': "27.5WP", '38WP': "28WP", '39WP': "28.5WP", '40WP': "29.5WP", '41WP': "30WP",
    '42WP': "30.5WP", '43WP': "31WP", '44WP': "32WP", '45WP': "32.5WP", '46WP': "33WP", '47WP': "34WP",
    '48WP': "34.5WP",
    '49WP': "35WP", '50WP': "36WP", '51WP': "36.5WP", '52WP': "37WP", '53WP': "37.5WP", '54WP': "38WP",
    '55WP': "38.5WP",
    '56WP': "39WP", '57WP': "40WP", '58WP': "40.5WP", '59WP': "41WP", '60WP': "41.5WP", '61WP': "42WP",
    '62WP': "42.5WP",
    '63WP': "43WP", '64WP': "43.5WP", '65WP': "44WP", '66WP': "44WP", '67WP': "44.5WP", '68WP': "45WP",
    '69WP': "45.5WP", '70WP': "46WP", '71WP': "46WP", '72WP': "46.5WP", '73WP': "46.5WP", '74WP': "47WP",
    '75WP': "47.5WP", '1WE': "1WE",
    '2WE': "2WE", '3WE': "3WE", '4WE': "4WE", '5WE': "6WE", '6WE': "7WE", '7WE': "8WE", '8WE': "9WE", '9WE': "10WE",
    '10WE': "11WE",
    '11WE': "13WE", '12WE': "14WE", '13WE': "15WE", '14WE': "16WE", '15WE': "17WE", '16WE': "18WE", '17WE': "19WE",
    '18WE': "20WE",
    '19WE': "22WE", '20WE': "23WE", '21WE': "24WE", '22WE': "25WE", '23WE': "26WE", '24WE': "27WE", '25WE': "28WE",
    '26WE': "29WE", '27WE': "30WE", '28WE': "32WE", '29WE': "33WE", '30WE': "34WE", '31WE': "35WE", '32WE': "36WE",
    '33WE': "37WE",
    '34WE': "38WE", '35WE': "39WE", '36WE': "41WE", '37WE': "42WE", '38WE': "43WE", '39WE': "44WE", '40WE': "45WE",
    '41WE': "46WE",
    '42WE': "47WE", '43WE': "48WE", '44WE': "50WE", '45WE': "51WE", '46WE': "52WE", '47WE': "53WE", '48WE': "54WE",
    '49WE': "55WE",
    '50WE': "56WE", '51WE': "58WE", '52WE': "59WE", '53WE': "60WE", '54WE': "61WE", '55WE': "62WE", '56WE': "63WE",
    '57WE': "64WE",
    '58WE': "65WE"
}


def correct_sheet(file_path):
    print("Converting the A1 values to corrected A1 values")
    # Open the Excel file
    wb = openpyxl.load_workbook(file_path)

    # Iterate over the sheet names in the workbook
    for sheet_name in wb.sheetnames:
        # Select the sheet you want to work with
        sheet = wb[sheet_name]

        # Iterate over the cells in the column you want to read
        for cell in sheet['C']:
            # Get the value of the cell
            old_value = cell.value
            # Convert the value using the dictionary
            new_value = conversion_dict.get(old_value, old_value)
            # Write the new value to the cell in the other column
            cell.offset(column=1).value = new_value

        for i, cell in enumerate(sheet['D']):
            if i == 0:
                continue
            # Extract the angle value and label from the cell value
            a4_value = cell.value[0:-2]  # Extract the angle value string
            a4_value = float(a4_value)
            label = cell.value[-2:]

            # Calculate the angle based on the label
            if label == "WE":
                r_value = 65 - a4_value
            elif label == "WP":
                r_value = a4_value
            elif label == "EP":
                r_value = a4_value
            else:
                r_value = 65 - a4_value

            # Write the calculated angle to the "angles" column
            sheet.cell(row=i + 1, column=24).value = r_value

        for i, cell in enumerate(sheet['B']):
            # Skip the first row
            if i == 0:
                continue

            # Extract the A1 value from the cell
            a1_value = cell.value
            # Extract the label from the corresponding cell in column D
            label = sheet.cell(row=i + 1, column=4).value[-2:]

            # Calculate the "theta" value based on the angle value and label
            if label == "WE":
                if 0 <= a1_value < 90:
                    theta_value = a1_value + 270
                else:
                    theta_value = a1_value - 90
            elif label == "WP":
                if 0 <= a1_value < 90:
                    theta_value = a1_value + 270
                else:
                    theta_value = a1_value - 90
            elif label == "EP":
                theta_value = 270 - a1_value
            else:
                theta_value = 270 - a1_value

            # Write the calculated "theta" value to the "theta" column
            sheet.cell(row=i + 1, column=25).value = theta_value

        for i, cell in enumerate(sheet['Y']):
            # Skip the first row
            if i == 0:
                continue
            # Extract the A1 value from the cell
            degree_value = cell.value
            # Calculate the "radians" value based on the degree value
            radian_value = math.radians(degree_value)
            # Write the calculated "radian" value to the "radian" column
            sheet.cell(row=i + 1, column=26).value = radian_value

        for i, (radian_value, r_value) in enumerate(zip(sheet['Z'], sheet['X'])):
            # Skip the first row
            if i == 0:
                continue
            r_value = r_value.value / 6.5
            radian_value = radian_value.value
            x_value = r_value * math.cos(radian_value)
            y_value = r_value * math.sin(radian_value)
            # Write the calculated "radian" value to the "radian" column
            sheet.cell(row=i + 1, column=27).value = x_value
            sheet.cell(row=i + 1, column=28).value = y_value

        for i, cell in enumerate(sheet['G']):
            # Skip the first row
            if i == 0:
                continue
            original_crystal_size = cell.value
            sheet.cell(row=i + 1, column=7).value = round((original_crystal_size * 1.25), 2)

        # Set the value of the heading cell for the new column
        sheet.cell(row=1, column=4).value = "Corrected A4 values"
        sheet.cell(row=1, column=24).value = "r values"
        sheet.cell(row=1, column=25).value = "theta values"
        sheet.cell(row=1, column=26).value = "radian values"
        sheet.cell(row=1, column=27).value = "x cartesian values"
        sheet.cell(row=1, column=28).value = "y cartesian values"

    # Save the changes to the file
    wb.save(file_path)
