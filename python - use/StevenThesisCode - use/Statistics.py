# this file is going to define 2 functions that will go through the excel sheet and create a degree of orientation
# for each of the thin section samples, along with the spherical aperture of the schmidt diagrams
import math
import pandas as pd
import numpy as np
import openpyxl
import matplotlib.pyplot as plt


def calculate_orientation(filepath):
    # read the excel sheet
    wb = openpyxl.load_workbook(filepath)
    # calculate vector components for each c-axis and add them up
    for sheet_name in wb.sheetnames:
        # Select the sheet you want to work with
        sheet = wb[sheet_name]
        n = 0
        x_sum = 0
        y_sum = 0
        z_sum = 0

        # Iterate over the cells in the column you want to read

        for i, cell in enumerate(sheet['D']):
            if i == 0:
                continue
            # Extract the angle value and label from the cell value
            a4_value = cell.value[0:-2]  # Extract the angle value string
            a4_value = float(a4_value)
            label = (cell.value[-2:]).upper()

            if label == "WE" or label == "WP":
                west = True
            elif label == "EP" or label == "EE":
                west = False
            else:
                raise ValueError("Invalid label in row " + str(i) + "")

            if west:
                a4_value = a4_value * -1
            phi = math.radians(a4_value)
            theta_cell = sheet.cell(row=cell.row, column=2)
            theta_value = theta_cell.value
            theta = math.radians(theta_value)
            x_sum += math.sin(theta) * math.cos(phi)
            y_sum += math.sin(theta) * math.sin(phi)
            z_sum += math.cos(theta)
            n += 1

        # calculate vector sum and length
        vector_sum = (x_sum, y_sum, z_sum)
        vector_length = math.sqrt(x_sum ** 2 + y_sum ** 2 + z_sum ** 2)
        # calculate degree of orientation
        R_percent = ((2 * abs(vector_length) - n) / n) * 100
        R_percent = round(R_percent, 2)

        # write degree of orientation to excel sheet
        sheet.cell(row=1, column=30).value = "Degree of Orientation"
        sheet.cell(row=2, column=30).value = R_percent

    wb.save(filepath)


def spherical_aperture(filepath):
    wb = openpyxl.load_workbook(filepath)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        total_x = 0
        total_y = 0
        total_n = 0
        for i, cell in enumerate(sheet['AA']):
            if i == 0:
                continue
            x_cartesian_value = cell.value
            total_x += x_cartesian_value
            y_cartesian_value = sheet.cell(row=cell.row, column=28).value
            total_y += y_cartesian_value
            total_n += 1

        average_x = total_x / total_n
        average_y = total_y / total_n
        sheet.cell(row=1, column=29).value = "Distances to average point"
        distance_list = []

        for i, cell in enumerate(sheet['AA']):
            if i == 0:
                continue
            x_cartesian_value = cell.value
            y_cartesian_value = sheet.cell(row=cell.row, column=28).value
            x_difference = x_cartesian_value - average_x
            y_difference = y_cartesian_value - average_y
            distance = math.sqrt(x_difference ** 2 + y_difference ** 2)
            distance = round(distance, 2)
            sheet.cell(row=cell.row, column=29).value = distance
            distance_list.append(distance)

        nintyth_percentile = round(total_n * 0.9)
        distance_list.sort()
        ninety_percentile_distance = distance_list[nintyth_percentile]
        sa = (ninety_percentile_distance / 10) * 100
        sa = round(sa, 2)
        sheet.cell(row=1, column=31).value = "Spherical Aperture"
        sheet.cell(row=2, column=31).value = sa

    wb.save(filepath)


def analyze_crystal_sizes(file_path):
    """
    This function takes the file path of an Excel workbook containing multiple sheets,
    each with a list of crystal sizes in rows. It returns a dictionary where each key
    is the sheet name and each value is a dictionary containing the mean, median, mode,
    range, and standard deviation of the crystal sizes in that sheet.
    """

    # Load the Excel workbook into a Pandas dataframe
    data = pd.read_excel(file_path, sheet_name=None, header=None)

    # Create a list to store the results
    results = []

    # Iterate through each sheet in the workbook
    for sheet_name, sheet_data in data.items():
        # Extract the crystal sizes from row G
        crystal_sizes = sheet_data.iloc[6:, 6].values

        # Convert the crystal sizes to numeric data
        crystal_sizes = pd.to_numeric(crystal_sizes, errors='coerce')

        # Remove any rows that contain non-numeric data
        crystal_sizes = pd.Series(crystal_sizes).dropna()

        # Calculate the statistical measures
        mean = round(crystal_sizes.mean(), 2)
        median = round(crystal_sizes.median(), 2)
        mode = round(crystal_sizes.mode()[0], 2) if not crystal_sizes.mode().empty else None
        range_val = round(crystal_sizes.max() - crystal_sizes.min(), 2)
        stdev = round(crystal_sizes.std(), 2)
        number_of_crystals = round(len(crystal_sizes))

        # Store the results in the dictionary
        result = {'sample': sheet_name, 'mean': mean, 'median': median, 'mode': mode,
                  'range': range_val, 'stdev': stdev, 'number_of_crystals': number_of_crystals}

        results.append(result)

    # Create a DataFrame from the results list
    df = pd.DataFrame(results)

    # Set the sample column as the index of the DataFrame
    df.set_index('sample', inplace=True)

    # Create a table from the DataFrame
    fig, ax = plt.subplots()
    ax.axis('off')
    ax.axis('tight')
    ax.table(cellText=df.values, colLabels=df.columns, rowLabels=df.index, loc='center')
    plt.savefig('table_image.png', dpi=300)

    # Create a box plot of the crystal sizes for each sample
    fig, ax = plt.subplots()
    for sheet_name, sheet_data in data.items():
        crystal_sizes = sheet_data.iloc[6:, 6].values
        crystal_sizes = pd.to_numeric(crystal_sizes, errors='coerce')
        crystal_sizes = pd.Series(crystal_sizes).dropna()
        ax.boxplot(crystal_sizes, positions=[df.index.get_loc(sheet_name)])
    ax.set_xticklabels(df.index, rotation=90)
    ax.set_ylabel('Crystal sizes')
    plt.tight_layout()
    plt.savefig('boxplot_outliers.png', dpi=300)

    # Create a box plot of the crystal sizes for each sample - excluding outliers
    fig, ax = plt.subplots()
    for sheet_name, sheet_data in data.items():
        crystal_sizes = sheet_data.iloc[6:, 6].values
        crystal_sizes = pd.to_numeric(crystal_sizes, errors='coerce')
        crystal_sizes = pd.Series(crystal_sizes).dropna()
        ax.boxplot(crystal_sizes, positions=[df.index.get_loc(sheet_name)], showfliers=False)
    ax.set_xticklabels(df.index, rotation=90)
    ax.set_ylabel('Crystal sizes')
    plt.tight_layout()
    plt.savefig('boxplot_no_outliers.png', dpi=300)

    return df


def visualisation(filepath):
    wb = openpyxl.load_workbook(filepath)

    r_values_all = []
    s_values_all = []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for i, cell in enumerate(sheet['AD']):
            if i == 0:
                continue
            elif i > 1:
                break
            r_value = cell.value
            s_value = sheet.cell(row=cell.row, column=31).value
            r_values_all.append(r_value)
            s_values_all.append(s_value)

    # Create a line graph of the R orientation and spherical aperture values from all sheets
    fig, ax = plt.subplots(figsize=(8, 6))
    ax.plot(r_values_all, label='R orientation', marker='o', linewidth=2, markersize=8)
    ax.plot(s_values_all, label='spherical aperture', marker='s', linewidth=2, markersize=8)
    ax.legend(fontsize=12, loc='center left')
    ax.set_xlabel('Sample Number', fontsize=14)
    ax.set_ylabel('Value', fontsize=14)
    ax.set_title('Comparison of R orientation and spherical aperture', fontsize=16)
    ax.tick_params(axis='both', which='major', labelsize=12)
    plt.tight_layout()
    plt.savefig('r vs sa', dpi=300)
