import openpyxl


def correct_sheet(file_path, conversion_dict):
    print("Converting the A1 values to corrected A1 values")
    # Open the Excel file
    wb = openpyxl.load_workbook(file_path)

    # Iterate over the sheet names in the workbook
    for sheet_name in wb.sheetnames:
        # Select the sheet you want to work with
        sheet = wb[sheet_name]

        # Iterate over the cells in the column you want to read
        for cell in sheet['C']:
            # Get the value of the cell
            old_value = cell.value
            # Convert the value using the dictionary
            new_value = conversion_dict.get(old_value, old_value)
            # Write the new value to the cell in the other column
            cell.offset(column=1).value = new_value

        for i, cell in enumerate(sheet['D']):
            if i == 0:
                continue
            # Extract the angle value and label from the cell value
            a4_value = cell.value[0:-2]  # Extract the angle value string
            a4_value = float(a4_value)
            label = cell.value[-2:]

            # Calculate the angle based on the label
            if label == "WE":
                r_value = 65 - a4_value
            elif label == "WP":
                r_value = a4_value
            elif label == "EP":
                r_value = a4_value
            else:
                r_value = 65 - a4_value

            # Write the calculated angle to the "angles" column
            sheet.cell(row=i + 1, column=24).value = r_value

        for i, cell in enumerate(sheet['B']):
            # Skip the first row
            if i == 0:
                continue

            # Extract the A1 value from the cell
            a1_value = cell.value
            # Extract the label from the corresponding cell in column D
            label = sheet.cell(row=i + 1, column=4).value[-2:]

            # Calculate the "theta" value based on the angle value and label
            if label == "WE":
                if 0 <= a1_value < 90:
                    theta_value = a1_value + 270
                else:
                    theta_value = a1_value - 90
            elif label == "WP":
                if 0 <= a1_value < 90:
                    theta_value = a1_value + 270
                else:
                    theta_value = a1_value - 90
            elif label == "EP":
                theta_value = 270 - a1_value
            else:
                theta_value = 270 - a1_value

            # Write the calculated "theta" value to the "theta" column
            sheet.cell(row=i + 1, column=25).value = theta_value

        # Set the value of the heading cell for the new column
        sheet.cell(row=1, column=4).value = "Corrected A4 values"
        sheet.cell(row=1, column=24).value = "r values"
        sheet.cell(row=1, column=25).value = "theta values"

    # Save the changes to the file
    wb.save(file_path)
