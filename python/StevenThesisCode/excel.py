import openpyxl
import os


class Excel:
    def __init__(self, folder_path, output_filename):
        self.folder_path = folder_path
        self.output_filename = output_filename
        self.files = os.listdir(self.folder_path)
        self.workbook = openpyxl.Workbook()
        self.workbook.remove_sheet(self.workbook.active)

    def combine_sheets(self):
        print("Combining sheets from Excel files...")
        desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')
        file_path = os.path.join(desktop_path, self.output_filename)
        if os.path.exists(file_path):
            # File already exists, handle it appropriately
            # For example, you can prompt the user to choose a different file name
            # or automatically generate a unique file name
            pass
        else:
            for file in self.files:
                if file.endswith('.xlsx'):
                    file_name, file_ext = os.path.splitext(file)
                    wb = openpyxl.load_workbook(os.path.join(self.folder_path, file))
                    for sheet in wb:
                        new_sheet = self.workbook.create_sheet(title=file_name)
                        new_sheet.insert_rows(0)
                        for row in sheet.rows:
                            for cell in row:
                                new_sheet[cell.coordinate].value = cell.value
            self.workbook.save(file_path)

    def update_values(self, key, new_value):
        sheets = self.workbook.sheetnames
        for sheet in sheets:
            current_sheet = self.workbook[sheet]
            for row in current_sheet.iter_rows():
                cell_value = row[0].value
                if cell_value == key:
                    row[1].value = new_value
        self.workbook.save(self.output_filename)
