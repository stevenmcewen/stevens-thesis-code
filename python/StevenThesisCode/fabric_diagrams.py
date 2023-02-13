import openpyxl
import math
from collections import defaultdict
import matplotlib.pyplot as plt
from scipy.spatial import Delaunay
import numpy as np
from matplotlib.tri import Triangulation
from matplotlib.colors import ListedColormap
import matplotlib.patches as mpatches


def make_fabric_diagrams(filepath_data_collecting, images_file_location):
    wb = openpyxl.load_workbook(filepath_data_collecting)

    # Iterate over the sheet names in the workbook
    for sheet_name in wb.sheetnames:
        # Select the sheet you want to work with
        sheet = wb[sheet_name]

        number_of_readings = sheet['E2'].value
        date = sheet['H2'].value
        x_coordinate = sheet['J2'].value
        y_coordinate = sheet['K2'].value
        sample_number = sheet['I2'].value

        grid = defaultdict(int)
        for x in range(-10, 11):
            for y in range(-10, 11):
                for i, (x_value, y_value) in enumerate(zip(sheet['AA'], sheet['AB'])):
                    # Skip the first row
                    if i == 0:
                        continue
                    x_value = x_value.value
                    y_value = y_value.value
                    distance = math.sqrt((x - x_value) ** 2 + (y - y_value) ** 2)

                    if distance <= 1:
                        grid[(x, y)] += 1
                    else:
                        grid[(x, y)] += 0

                    tuples_with_z = [(x, y, z) for (x, y), z in grid.items()]

                    to_remove = {(-10, 10), (-9, 10), (-8, 10), (-7, 10), (-6, 10), (-5, 10), (-4, 10), (-3, 10),
                                 (-2, 10), (-1, 10),
                                 (10, 10), (9, 10), (8, 10), (7, 10), (6, 10), (5, 10), (4, 10), (3, 10), (2, 10),
                                 (1, 10),
                                 (-10, -10), (-9, -10), (-8, -10), (-7, -10), (-6, -10), (-5, -10), (-4, -10),
                                 (-3, -10), (-2, -10), (-1, -10),
                                 (10, -10), (9, -10), (8, -10), (7, -10), (6, -10), (5, -10), (4, -10), (3, -10),
                                 (2, -10),
                                 (1, -10), (-10, 9), (-9, 9), (-8, 9), (-7, 9), (-6, 9), (-5, 9), (10, 9), (9, 9),
                                 (8, 9), (7, 9), (6, 9), (5, 9)
                        , (-10, -9), (-9, -9), (-8, -9), (-7, -9), (-6, -9), (-5, -9), (10, -9), (9, -9), (8, -9),
                                 (7, -9), (6, -9), (5, -9),
                                 (-10, 8), (-9, 8), (-8, 8), (-7, 8), (10, 8), (9, 8), (8, 8), (7, 8), (-10, -8),
                                 (-9, -8), (-8, -8), (-7, -8), (10, -8), (9, -8), (8, -8), (7, -8),
                                 (-10, 7), (-9, 7), (-8, 7), (10, 7), (9, 7), (8, 7), (-10, -7), (-9, -7), (-8, -7),
                                 (10, -7), (9, -7), (8, -7),
                                 (-10, 6), (-9, 6), (10, 6), (9, 6), (-10, -6), (-9, -6), (10, -6), (9, -6),
                                 (-10, 5), (-9, 5), (10, 5), (9, 5), (-10, -5), (-9, -5), (10, -5), (9, -5),
                                 (-10, 4), (10, 4), (-10, -4), (10, -4), (-10, 3), (10, 3), (-10, -3), (10, -3),
                                 (-10, 2), (10, 2), (-10, -2), (10, -2), (-10, 1), (10, 1), (-10, -1), (10, -1)}
                    tuples_with_z = [tup for tup in tuples_with_z if (tup[0], tup[1]) not in to_remove]

                    polar_tuples = [(math.atan2(y, x), math.hypot(x, y), z) for (x, y, z) in tuples_with_z]

        fig = plt.figure()
        ax = fig.add_subplot(111, projection='polar')

        angles = [tup[0] for tup in polar_tuples]
        distances = [tup[1] for tup in polar_tuples]
        z_values = [tup[2] for tup in polar_tuples]
        number_of_rows = sheet.max_row - 1
        z_values = [z_value/number_of_rows for z_value in z_values]

        levels = [0.01, 0.02, 0.03, 0.04, 0.05, 0.06, 0.07, 0.08, 0.09, 0.1, 0.11, 0.12, 0.13, 0.14, 0.15, 0.16, 0.17]
        colors = ['#FFF3E0', '#FFE0B2', '#FFCC80', '#FFB74D', '#FFA726', '#FF9800', '#FB8C00', '#F57C00', '#EF6C00', '#E65100', '#D84315', '#BF360C', '#A93226', '#8B0000', '#7F0000', '#660000', '#330000']
        cmap = ListedColormap(colors)
        legend_text = f"Number of readings: {number_of_readings}\nDate: {date}\nX-coordinate: {x_coordinate}\nY-coordinate: {y_coordinate}\nSample number: {sample_number}"
        legend_ax = fig.add_axes([0.025, 0.025, 0.325, 0.325])
        legend_ax.axis('off')
        legend_ax.text(0, 0, legend_text, va='bottom', fontsize=9)

        triangulation = Delaunay(np.column_stack((angles, distances)), incremental=True)
        triangulation = Triangulation(angles, distances, triangles=triangulation.simplices)
        sm = plt.cm.ScalarMappable(cmap=cmap, norm=plt.Normalize(vmin=min(z_values), vmax=0.17))
        sm.set_array(z_values)
        triangulation = ax.tricontour(triangulation, z_values, cmap=cmap,
                                      norm=plt.Normalize(vmin=min(z_values), vmax=0.17))
        plt.colorbar(mappable=sm, ax=ax, orientation='vertical', aspect=10)
        ax.get_xaxis().set_ticklabels([])
        ax.get_yaxis().set_ticklabels([])
        ax.set_title(sheet_name)
        ax.grid(False)
        plt.grid(visible=False)
        plt.savefig(f"{images_file_location}{sheet_name}.png", dpi=300)


